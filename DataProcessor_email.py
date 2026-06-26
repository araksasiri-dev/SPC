# DataProcessor_email.py
import re
import os
from openpyxl import load_workbook
from database import DatabaseManager
from gmail_oauth import GmailOAuth
from dotenv import load_dotenv

load_dotenv()

class DataProcessor:
    def __init__(self, excel_path="users_dup.xlsx"):
        self.excel_path = excel_path
        self.db = DatabaseManager()
        self.gmail = None
        self.raw_data = []
        self.clean_data = []
        self.duplicate_email = []
        self.duplicate_phone = []
        
        # ✅ เปลี่ยนให้อ่านจาก C:\secrets\
        self.credentials_file = os.getenv("GMAIL_CREDENTIALS", r"C:\secrets\credentials.json")
        self.token_file = os.getenv("GMAIL_TOKEN", r"C:\secrets\token.pickle")
        self.recipient_email = os.getenv("RECIPIENT_EMAIL", "ar0816250183@gmail.com")
        
        # ✅ ตรวจสอบว่าไฟล์มีอยู่จริง
        if not os.path.exists(self.credentials_file):
            print(f"⚠️ ไม่พบไฟล์ credentials.json ที่ {self.credentials_file}")
        if not os.path.exists(self.token_file):
            print(f"⚠️ ไม่พบไฟล์ token.pickle ที่ {self.token_file}")
    
    # ========== Utility ==========
    def _safe_str(self, value):
        if value is None:
            return ""
        return str(value).strip()
    
    # ========== Excel ==========
    def read_excel(self):
        wb = load_workbook(self.excel_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if not headers or all(h is None for h in headers):
            wb.close()
            return []
        
        self.raw_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                break
            user_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    value = row[i]
                    if header == "phone" and value is not None:
                        value = str(value).zfill(10)
                    user_dict[header] = value
            self.raw_data.append(user_dict)
        
        wb.close()
        print(f"✅ อ่านข้อมูลจาก {self.excel_path} สำเร็จ ({len(self.raw_data)} รายการ)")
        return self.raw_data
    
    def write_results_to_excel(self, results, output_path="results_gmail.xlsx"):
        wb = load_workbook(self.excel_path)
        ws = wb.active
        col_status = ws.max_column + 1
        ws.cell(row=1, column=col_status, value="STATUS")
        
        status_map = {}
        for item in results:
            username = item.get("username")
            status = item.get("status")
            if username:
                status_map[username] = status
        
        for row_num in range(2, ws.max_row + 1):
            username_cell = ws.cell(row=row_num, column=1)
            username = username_cell.value
            
            if username in status_map:
                ws.cell(row=row_num, column=col_status, value=status_map[username])
            else:
                ws.cell(row=row_num, column=col_status, value="SKIPPED (duplicate)")
        
        wb.save(output_path)
        wb.close()
        print(f"✅ บันทึกผลลัพธ์ลง {output_path} เรียบร้อย")
    
    # ========== Data Validation ==========
    def process_data(self):
            if not self.raw_data:
                self.read_excel()
            
            email_seen = {}
            phone_seen = {}
            self.clean_data = []
            self.duplicate_email = []
            self.duplicate_phone = []
            
            for user in self.raw_data:
                email = self._safe_str(user.get("email", "")).strip()
                phone = self._safe_str(user.get("phone", "")).strip()
                
                # ตรวจสอบกรณีข้อมูลว่าง
                if not email or not phone:
                    continue
                
                # 🟢 เปลี่ยนตรรกะ: เช็กสถานะแยกขาดจากกัน ไม่ตัดตอนด้วย continue มั่วซั่ว
                is_email_dup = email in email_seen
                is_phone_dup = phone in phone_seen
                
                if is_email_dup:
                    self.duplicate_email.append(user)
                    
                if is_phone_dup:
                    self.duplicate_phone.append(user)
                
                # 🟢 บันทึกประวัติการพบเจอทันทีเพื่อใช้เช็กในแถวถัดไป
                if email: email_seen[email] = True
                if phone: phone_seen[phone] = True
                
                # 🟢 ข้อมูลจะถือว่า Clean ก็ต่อเมื่อ "ไม่ซ้ำทั้งอีเมลและเบอร์โทร"
                if not is_email_dup and not is_phone_dup:
                    self.clean_data.append(user)
            
            print(f"✅ ตรวจสอบข้อมูลสำเร็จ: ผ่าน {len(self.clean_data)} รายการ, อีเมลซ้ำ {len(self.duplicate_email)} รายการ, เบอร์โทรซ้ำ {len(self.duplicate_phone)} รายการ")
            
            return {
                "clean_data": self.clean_data,
                "duplicate_email": self.duplicate_email,
                "duplicate_phone": self.duplicate_phone
            }
            
    def get_clean_users(self):
        return self.clean_data
    
    def get_duplicate_email_users(self):
        return self.duplicate_email
    
    def get_duplicate_phone_users(self):
        return self.duplicate_phone
    
    def validate_email(self, email):
        email_str = self._safe_str(email)
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email_str) is not None
    
    # ========== Database ==========
    def save_results_to_db(self, results):
        for item in results:
            self.db.insert_user(
                username=item.get("username"),
                email=item.get("email"),
                phone=item.get("phone"),
                status=item.get("status"),
                error_message=item.get("error_message", "")
            )
        summary = self.db.get_summary()
        print(f"✅ บันทึก Database สำเร็จ: {summary}")
        return summary
    
    def get_db_summary(self):
        return self.db.get_summary()
    
    def get_all_users_from_db(self):
        return self.db.get_all_users()
    
    def clear_database(self):
        self.db.clear_all()
        print("🗑️ ล้าง Database เรียบร้อย")
    
    # ========== Gmail OAuth ==========
    def init_gmail(self, credentials_file=None, token_file=None):
        if credentials_file is None:
            credentials_file = self.credentials_file
        if token_file is None:
            token_file = self.token_file
        
        if self.gmail is None:
            try:
                self.gmail = GmailOAuth(credentials_file, token_file)
                print("✅ Gmail OAuth พร้อมใช้งาน")
            except Exception as e:
                print(f"❌ เริ่มต้น Gmail OAuth ล้มเหลว: {e}")
                return None
        
        return self.gmail
    
    def send_email_report(self, recipient_email, summary, user_list, duplicate_report=""):
        """
        ส่งรายงานทางอีเมลด้วย Gmail OAuth
        
        Args:
            recipient_email (str): อีเมลผู้รับ
            summary (dict): สรุปสถิติ
            user_list (list): รายละเอียดผู้ใช้
            duplicate_report (str): รายงานข้อมูลซ้ำ (optional)
        
        Returns:
            bool: True ถ้าส่งสำเร็จ
        """
        if self.gmail is None:
            self.init_gmail()
        
        if self.gmail is None:
            print("❌ Gmail OAuth ไม่พร้อมใช้งาน")
            return False
        
        try:
            # ส่งไปยัง gmail_oauth.py
            result = self.gmail.send_registration_summary(
                recipient_email, summary, user_list, duplicate_report
            )
            return result
        except Exception as e:
            print(f"❌ ส่งอีเมลล้มเหลว: {e}")
            return False
    
    def send_test_email(self, recipient_email=None):
        if recipient_email is None:
            recipient_email = self.recipient_email
        
        if self.gmail is None:
            self.init_gmail()
        
        if self.gmail is None:
            print("❌ Gmail OAuth ไม่พร้อมใช้งาน")
            return False
        
        return self.gmail.send_email(
            recipient_email=recipient_email,
            subject="✅ ทดสอบ Gmail OAuth",
            body="สวัสดี! การทดสอบ Gmail OAuth ทำงานสำเร็จ"
        )
    
    def get_duplicate_report(self):
        """
        สร้างรายงานข้อมูลซ้ำ พร้อมเลขลำดับและขึ้นบรรทัดใหม่
        """
        report = ""
        
        if self.duplicate_email:
            report += "📧 อีเมลซ้ำ ({} ราย):\n".format(len(self.duplicate_email))
            for idx, user in enumerate(self.duplicate_email, start=1):
                report += "   {}. {} ({})\n".format(idx, user.get('username'), user.get('email'))
        else:
            report += "✅ ไม่มีอีเมลซ้ำ\n"
        
        report += "\n"  # บรรทัดว่างระหว่างส่วน
        
        if self.duplicate_phone:
            report += "📱 เบอร์โทรซ้ำ ({} ราย):\n".format(len(self.duplicate_phone))
            for idx, user in enumerate(self.duplicate_phone, start=1):
                report += "   {}. {} ({})\n".format(idx, user.get('username'), user.get('phone'))
        else:
            report += "✅ ไม่มีเบอร์โทรซ้ำ\n"
        
        return report
    
    # ========== Summary ==========
    def get_full_summary(self):
        return {
            "excel": {
                "total": len(self.raw_data),
                "clean": len(self.clean_data),
                "duplicate_email": len(self.duplicate_email),
                "duplicate_phone": len(self.duplicate_phone)
            },
            "database": self.db.get_summary()
        }
        
 
    def save_all_results_to_db(self, db_results, dup_emails, dup_phones):
        import sqlite3
        import os
        
        # ล็อกพิกัดโฟลเดอร์ปลอดภัยในโปรเจกต์ (รองรับ Windows 100%)
        current_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(current_dir, "users.db")
        
        print(f"\n[SYSTEM-CHECK] โค้ดทำงานแล้ว! ระบบกำลังบันทึกข้อมูลแบบสะสมไปที่: {db_path}")
            
        local_conn = sqlite3.connect(db_path)
        local_cursor = local_conn.cursor()
        
        # 🟢 เอา DROP TABLE ออกไปแล้ว! 
        # 🟢 เปลี่ยนมาใช้ CREATE TABLE IF NOT EXISTS และไม่มีกฎ UNIQUE ค้ำคอ
        # เพื่อเปิดทางให้ข้อมูลในรอบถัดๆ ไปไหลเข้าต่อท้าย (Append) ได้เรื่อยๆ โดยไม่ทำลายข้อมูลเก่า
        local_cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                email TEXT,
                phone TEXT,
                status TEXT,
                registered_at DATETIME DEFAULT (datetime('now', 'localtime')),
                error_message TEXT
            )
        """)
        
        # 1. บันทึกกลุ่มที่ผ่านการกรอกหน้าเว็บ 
        for user in db_results:
            username = user.get('username', '') if hasattr(user, 'get') else user['username']
            email = user.get('email', '') if hasattr(user, 'get') else user['email']
            phone = user.get('phone', '') if hasattr(user, 'get') else user['phone']
            status = user.get('status', '') if hasattr(user, 'get') else user['status']
            
            local_cursor.execute("""
                INSERT INTO users (username, email, phone, status) 
                VALUES (?, ?, ?, ?)
            """, (username, email, phone, status))
            
        # 2. บันทึกกลุ่มอีเมลซ้ำ (บันทึกตรงๆ เพื่อเก็บประวัติ Audit Log)
        for user in dup_emails:
            username = user.get('username', '') if hasattr(user, 'get') else user['username']
            email = user.get('email', '') if hasattr(user, 'get') else user['email']
            phone = user.get('phone', '') if hasattr(user, 'get') else user['phone']
            
            local_cursor.execute("""
                INSERT INTO users (username, email, phone, status) 
                VALUES (?, ?, ?, ?)
            """, (username, email, phone, 'SKIPPED_DUP_EMAIL'))

        # 3. บันทึกกลุ่มเบอร์โทรซ้ำ
        for user in dup_phones:
            username = user.get('username', '') if hasattr(user, 'get') else user['username']
            email = user.get('email', '') if hasattr(user, 'get') else user['email']
            phone = user.get('phone', '') if hasattr(user, 'get') else user['phone']
            
            local_cursor.execute("""
                INSERT INTO users (username, email, phone, status) 
                VALUES (?, ?, ?, ?)
            """, (username, email, phone, 'SKIPPED_DUP_PHONE'))
            
        local_conn.commit()
        local_conn.close()
        
        total_inserted = len(db_results) + len(dup_emails) + len(dup_phones)
        print(f"[SYSTEM-CHECK] 🟢 บันทึกข้อมูลแบบสะสมสำเร็จในรอบนี้: {total_inserted} แถว")
        return total_inserted
 
    def get_full_summary_with_duplicates(self):
        """
        ดึงสรุปข้อมูลทั้งหมด (รวมข้อมูลซ้ำ) - แก้ไขให้ส่งค่า duplicate_email และ duplicate_phone
        """
        # ดึงข้อมูลจาก Database
        db_summary = self.db.get_summary()
        
        # ข้อมูลจาก process_data
        total_clean = len(self.clean_data)
        total_dup_email = len(self.duplicate_email)
        total_dup_phone = len(self.duplicate_phone)
        total_all = len(self.raw_data)
        
        success = db_summary.get("success", 0)
        failed = db_summary.get("failed", 0)
        skipped = total_dup_email + total_dup_phone
        
        # ตรวจสอบว่า total = success + failed + skipped หรือไม่
        calculated_total = success + failed + skipped
        if total_all != calculated_total:
            print(f"⚠️ ตัวเลขไม่ตรงกัน: total={total_all}, calculated={calculated_total}")
            # ปรับ total ให้ตรงกับ success + failed + skipped
            total_all = calculated_total
        
        return {
            "total": total_all,
            "clean": total_clean,
            "duplicate_email": total_dup_email,   # ✅ ส่งค่าอีเมลซ้ำ
            "duplicate_phone": total_dup_phone,   # ✅ ส่งค่าเบอร์โทรซ้ำ
            "success": success,
            "failed": failed,
            "skipped": skipped
        }
 

    def get_summary_from_excel(self, excel_path="results_gmail.xlsx"):
        """
        อ่านข้อมูลจากไฟล์ Excel และสรุปผล
        Returns:
            dict: {"total": int, "success": int, "failed": int, "skipped": int}
        """
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            total = 0
            success = 0
            failed = 0
            skipped = 0
            
            # อ่านข้อมูลตั้งแต่แถวที่ 2 เป็นต้นไป
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    break
                
                total += 1
                status = row[3] if len(row) > 3 else ""  # STATUS อยู่คอลัมน์ D (index 3)
                
                if status == "SUCCESS":
                    success += 1
                elif status and "SKIPPED" in str(status):
                    skipped += 1
                elif status == "FAIL":
                    failed += 1
            
            wb.close()
            
            return {
                "total": total,
                "success": success,
                "failed": failed,
                "skipped": skipped
            }
        except Exception as e:
            print(f"❌ อ่านไฟล์ Excel ไม่สำเร็จ: {e}")
            return {"total": 0, "success": 0, "failed": 0, "skipped": 0}