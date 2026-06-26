# DataProcessor_email.py
import re
import os
from openpyxl import load_workbook
from database import DatabaseManager
from gmail_oauth import GmailOAuth
from dotenv import load_dotenv

load_dotenv()

class DataProcessor:
    def __init__(self, excel_path="users_dup.xlsx"):
        self.excel_path = excel_path
        self.db = DatabaseManager()
        self.gmail = None
        self.raw_data = []
        self.clean_data = []
        self.duplicate_email = []
        self.duplicate_phone = []
        
        self.credentials_file = os.getenv("GMAIL_CREDENTIALS", "credentials.json")
        self.token_file = os.getenv("GMAIL_TOKEN", "token.pickle")
        self.recipient_email = os.getenv("RECIPIENT_EMAIL", "admin@example.com")
    
    # ========== Utility ==========
    def _safe_str(self, value):
        if value is None:
            return ""
        return str(value).strip()
    
    # ========== Excel ==========
    def read_excel(self):
        wb = load_workbook(self.excel_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if not headers or all(h is None for h in headers):
            wb.close()
            return []
        
        self.raw_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                break
            user_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    value = row[i]
                    if header == "phone" and value is not None:
                        value = str(value).zfill(10)
                    user_dict[header] = value
            self.raw_data.append(user_dict)
        
        wb.close()
        print(f"✅ อ่านข้อมูลจาก {self.excel_path} สำเร็จ ({len(self.raw_data)} รายการ)")
        return self.raw_data
    
    def write_results_to_excel(self, results, output_path="results_gmail.xlsx"):
        wb = load_workbook(self.excel_path)
        ws = wb.active
        col_status = ws.max_column + 1
        ws.cell(row=1, column=col_status, value="STATUS")
        
        status_map = {}
        for item in results:
            username = item.get("username")
            status = item.get("status")
            if username:
                status_map[username] = status
        
        for row_num in range(2, ws.max_row + 1):
            username_cell = ws.cell(row=row_num, column=1)
            username = username_cell.value
            
            if username in status_map:
                ws.cell(row=row_num, column=col_status, value=status_map[username])
            else:
                ws.cell(row=row_num, column=col_status, value="SKIPPED (duplicate)")
        
        wb.save(output_path)
        wb.close()
        print(f"✅ บันทึกผลลัพธ์ลง {output_path} เรียบร้อย")
    
    # ========== Data Validation ==========
    def process_data(self):
        """ตรวจสอบข้อมูลซ้ำ แยกเป็น 3 ประเภท"""
        if not self.raw_data:
            self.read_excel()
        
        email_seen = {}
        phone_seen = {}
        
        self.clean_data = []
        self.duplicate_email = []
        self.duplicate_phone = []
        self.duplicate_both = []  # ✅ ใหม่: ซ้ำทั้ง Email และ Phone
        
        # ตรวจสอบข้อมูลซ้ำ
        for user in self.raw_data:
            email = self._safe_str(user.get("email", ""))
            phone = self._safe_str(user.get("phone", ""))
            username = self._safe_str(user.get("username", ""))
            
            if not email or not phone:
                continue
            
            # ตรวจสอบซ้ำ
            is_dup_email = email in email_seen
            is_dup_phone = phone in phone_seen
            
            if is_dup_email and is_dup_phone:
                # ✅ ซ้ำทั้ง Email และ Phone
                self.duplicate_both.append(user)
            elif is_dup_email:
                self.duplicate_email.append(user)
            elif is_dup_phone:
                self.duplicate_phone.append(user)
            else:
                self.clean_data.append(user)
            
            email_seen[email] = True
            phone_seen[phone] = True
        
        return {
            "clean_data": self.clean_data,
            "duplicate_email": self.duplicate_email,
            "duplicate_phone": self.duplicate_phone,
            "duplicate_both": self.duplicate_both   # ✅ ใหม่
        }
    
    def get_clean_users(self):
        return self.clean_data
    
    def get_duplicate_email_users(self):
        return self.duplicate_email
    
    def get_duplicate_phone_users(self):
        return self.duplicate_phone
    
    def validate_email(self, email):
        email_str = self._safe_str(email)
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email_str) is not None
    
    # ========== Database ==========
    def save_results_to_db(self, results):
        for item in results:
            self.db.insert_user(
                username=item.get("username"),
                email=item.get("email"),
                phone=item.get("phone"),
                status=item.get("status"),
                error_message=item.get("error_message", "")
            )
        summary = self.db.get_summary()
        print(f"✅ บันทึก Database สำเร็จ: {summary}")
        return summary
    
    def get_db_summary(self):
        return self.db.get_summary()
    
    def get_all_users_from_db(self):
        return self.db.get_all_users()
    
    def clear_database(self):
        self.db.clear_all()
        print("🗑️ ล้าง Database เรียบร้อย")
    
    # ========== Gmail OAuth ==========
    def init_gmail(self, credentials_file=None, token_file=None):
        if credentials_file is None:
            credentials_file = self.credentials_file
        if token_file is None:
            token_file = self.token_file
        
        if self.gmail is None:
            try:
                self.gmail = GmailOAuth(credentials_file, token_file)
                print("✅ Gmail OAuth พร้อมใช้งาน")
            except Exception as e:
                print(f"❌ เริ่มต้น Gmail OAuth ล้มเหลว: {e}")
                return None
        
        return self.gmail
    
    def send_email_report(self, recipient_email, summary, user_list, duplicate_report=""):
        """
        ส่งรายงานทางอีเมลด้วย Gmail OAuth
        
        Args:
            recipient_email (str): อีเมลผู้รับ
            summary (dict): สรุปสถิติ
            user_list (list): รายละเอียดผู้ใช้
            duplicate_report (str): รายงานข้อมูลซ้ำ (optional)
        
        Returns:
            bool: True ถ้าส่งสำเร็จ
        """
        if self.gmail is None:
            self.init_gmail()
        
        if self.gmail is None:
            print("❌ Gmail OAuth ไม่พร้อมใช้งาน")
            return False
        
        try:
            # ส่งไปยัง gmail_oauth.py
            result = self.gmail.send_registration_summary(
                recipient_email, summary, user_list, duplicate_report
            )
            return result
        except Exception as e:
            print(f"❌ ส่งอีเมลล้มเหลว: {e}")
            return False
    
    def send_test_email(self, recipient_email=None):
        if recipient_email is None:
            recipient_email = self.recipient_email
        
        if self.gmail is None:
            self.init_gmail()
        
        if self.gmail is None:
            print("❌ Gmail OAuth ไม่พร้อมใช้งาน")
            return False
        
        return self.gmail.send_email(
            recipient_email=recipient_email,
            subject="✅ ทดสอบ Gmail OAuth",
            body="สวัสดี! การทดสอบ Gmail OAuth ทำงานสำเร็จ"
        )
    
    def get_duplicate_report(self):
        """
        สร้างรายงานข้อมูลซ้ำ พร้อมเลขลำดับและแยกประเภท
        """
        report = ""
        
        if self.duplicate_email:
            report += "📧 อีเมลซ้ำ ({} ราย):\n".format(len(self.duplicate_email))
            for idx, user in enumerate(self.duplicate_email, start=1):
                report += "   {}. {} ({})\n".format(idx, user.get('username'), user.get('email'))
        else:
            report += "✅ ไม่มีอีเมลซ้ำ\n"
        
        report += "\n"
        
        if self.duplicate_phone:
            report += "📱 เบอร์โทรซ้ำ ({} ราย):\n".format(len(self.duplicate_phone))
            for idx, user in enumerate(self.duplicate_phone, start=1):
                report += "   {}. {} ({})\n".format(idx, user.get('username'), user.get('phone'))
        else:
            report += "✅ ไม่มีเบอร์โทรซ้ำ\n"
        
        report += "\n"
        
        # ✅ ใหม่: ซ้ำทั้งคู่
        if self.duplicate_both:
            report += "⚠️ ซ้ำทั้งอีเมลและเบอร์โทร ({} ราย):\n".format(len(self.duplicate_both))
            for idx, user in enumerate(self.duplicate_both, start=1):
                report += "   {}. {} (📧{} | 📱{})\n".format(
                    idx, 
                    user.get('username'), 
                    user.get('email'), 
                    user.get('phone')
                )
        else:
            report += "✅ ไม่มีข้อมูลซ้ำทั้งคู่\n"
        
        return report
    
    # ========== Summary ==========
    def get_full_summary(self):
        return {
            "excel": {
                "total": len(self.raw_data),
                "clean": len(self.clean_data),
                "duplicate_email": len(self.duplicate_email),
                "duplicate_phone": len(self.duplicate_phone)
            },
            "database": self.db.get_summary()
        }
 
    def save_all_results_to_db(self, clean_results, duplicate_email_results, 
                                duplicate_phone_results, duplicate_both_results):
        """บันทึกผลลัพธ์ทั้งหมดลง Database (ใช้ Logic เดียวกับ import_csv.py)"""
        
        # ✅ ใช้ Excel โดยตรง แทนการใช้ Logic ที่ซับซ้อน
        from openpyxl import load_workbook
        import sqlite3
        
        # 1. อ่านข้อมูลจาก Excel
        wb = load_workbook(self.excel_path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        # 2. เชื่อมต่อ Database
        conn = sqlite3.connect("users.db")
        cursor = conn.cursor()
        
        # 3. ล้างข้อมูลเก่า
        cursor.execute("DELETE FROM users")
        
        # 4. Insert ข้อมูลทั้งหมด
        success_count = 0
        skipped_count = 0
        email_seen = set()
        phone_seen = set()
        
        for row in rows:
            if not row or not row[0]:
                continue
            
            username = row[0]
            email = row[1] if len(row) > 1 else ""
            phone = row[2] if len(row) > 2 else ""
            
            # ✅ ใช้ Logic เดียวกับ import_csv.py
            if username.startswith('dup_'):
                status = "SKIPPED (Duplicate)"
                skipped_count += 1
            else:
                status = "SUCCESS"
                success_count += 1
            
            cursor.execute("""
                INSERT OR REPLACE INTO users 
                (username, email, phone, status, registered_at)
                VALUES (?, ?, ?, ?, datetime('now'))
            """, (username, email, phone, status))
        
        conn.commit()
        conn.close()
        
        summary = {
            "total": len(rows),
            "success": success_count,
            "failed": 0,
            "skipped": skipped_count
        }
        print(f"✅ บันทึก Database ทั้งหมด: {summary}")
        return summary
         

     
    def get_full_summary_with_duplicates(self):
            db_summary = self.db.get_summary()
            
            total_clean = len(self.clean_data)
            total_dup_email = len(self.duplicate_email)
            total_dup_phone = len(self.duplicate_phone)
            total_dup_both = len(self.duplicate_both)  # ✅ ต้องมี
            total_all = len(self.raw_data)
            
            success = db_summary.get("success", 0)
            failed = db_summary.get("failed", 0)
            skipped = total_dup_email + total_dup_phone + total_dup_both
            
            return {
                "total": total_all,
                "clean": total_clean,
                "duplicate_email": total_dup_email,
                "duplicate_phone": total_dup_phone,
                "duplicate_both": total_dup_both,  # ✅ ต้องมี
                "success": success,
                "failed": failed,
                "skipped": skipped
            }    
            
    def get_summary_from_excel(self, excel_path="results_gmail.xlsx"):
        """
        อ่านข้อมูลจากไฟล์ Excel และสรุปผล
        Returns:
            dict: {"total": int, "success": int, "failed": int, "skipped": int}
        """
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            total = 0
            success = 0
            failed = 0
            skipped = 0
            
            # อ่านข้อมูลตั้งแต่แถวที่ 2 เป็นต้นไป
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    break
                
                total += 1
                status = row[3] if len(row) > 3 else ""  # STATUS อยู่คอลัมน์ D (index 3)
                
                if status == "SUCCESS":
                    success += 1
                elif status and "SKIPPED" in str(status):
                    skipped += 1
                elif status == "FAIL":
                    failed += 1
            
            wb.close()
            
            return {
                "total": total,
                "success": success,
                "failed": failed,
                "skipped": skipped
            }
        except Exception as e:
            print(f"❌ อ่านไฟล์ Excel ไม่สำเร็จ: {e}")
            return {"total": 0, "success": 0, "failed": 0, "skipped": 0}