# force_sync_db.py
import sqlite3
import os
from openpyxl import load_workbook

def force_sync():
    excel_file = "users_dup.xlsx"
    
    # 1. ลบไฟล์เก่า
    if os.path.exists("users.db"):
        os.remove("users.db")
        print("🗑️ ลบ users.db เก่าแล้ว")
    
    # 2. สร้าง Database ใหม่
    conn = sqlite3.connect("users.db")
    cursor = conn.cursor()
    
    # 3. สร้างตาราง
    cursor.execute("""
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            email TEXT UNIQUE,
            phone TEXT UNIQUE,
            status TEXT,
            registered_at TEXT,
            error_message TEXT
        )
    """)
    
    # 4. อ่าน Excel
    wb = load_workbook(excel_file)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"📊 อ่านข้อมูลจาก {excel_file}: {len(rows)} ราย")
    
    # 5. Insert ข้อมูล
    success_count = 0
    dup_email_count = 0
    dup_phone_count = 0
    email_seen = set()
    phone_seen = set()
    
    for row in rows:
        if not row or not row[0]:
            continue
        
        username = row[0]
        email = row[1] if len(row) > 1 else ""
        phone = row[2] if len(row) > 2 else ""
        
        # ตรวจสอบสถานะ
        if username.startswith('dup_'):
            if email in email_seen and phone in phone_seen:
                status = "SKIPPED (Duplicate Both)"
            elif email in email_seen:
                status = "SKIPPED (Duplicate Email)"
                dup_email_count += 1
            elif phone in phone_seen:
                status = "SKIPPED (Duplicate Phone)"
                dup_phone_count += 1
            else:
                status = "SUCCESS"
                success_count += 1
        else:
            status = "SUCCESS"
            success_count += 1
        
        email_seen.add(email)
        phone_seen.add(phone)
        
        cursor.execute("""
            INSERT OR REPLACE INTO users 
            (username, email, phone, status, registered_at)
            VALUES (?, ?, ?, ?, datetime('now'))
        """, (username, email, phone, status))
    
    conn.commit()
    conn.close()
    
    print(f"\n📊 สรุปข้อมูลที่บันทึก:")
    print(f"   ✅ SUCCESS: {success_count} ราย")
    print(f"   📧 SKIPPED (Email): {dup_email_count} ราย")
    print(f"   📱 SKIPPED (Phone): {dup_phone_count} ราย")
    print(f"   {'='*40}")
    print(f"   รวมทั้งหมด: {success_count + dup_email_count + dup_phone_count} ราย")

if __name__ == "__main__":
    force_sync()