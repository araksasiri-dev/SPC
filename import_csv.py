# import_csv.py
import sqlite3
import csv
from openpyxl import load_workbook

def excel_to_csv():
    """แปลง Excel เป็น CSV ก่อน"""
    wb = load_workbook("users_dup.xlsx")
    ws = wb.active
    
    with open("users_temp.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(min_row=1, values_only=True):
            writer.writerow(row)
    print("✅ แปลง Excel → CSV สำเร็จ")

def import_to_sqlite():
    """นำเข้า CSV ไป SQLite"""
    conn = sqlite3.connect("users.db")
    cursor = conn.cursor()
    
    # สร้างตาราง
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            email TEXT,
            phone TEXT,
            status TEXT,
            registered_at TEXT,
            error_message TEXT
        )
    """)
    
    # ลบข้อมูลเก่า
    cursor.execute("DELETE FROM users")
    
    # อ่าน CSV และ Insert
    with open("users_temp.csv", "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)  # ข้าม header
        
        for row in reader:
            if not row or not row[0]:
                continue
            
            username = row[0]
            email = row[1] if len(row) > 1 else ""
            phone = row[2] if len(row) > 2 else ""
            
            # กำหนด status
            if username.startswith('dup_'):
                status = "SKIPPED (Duplicate)"
            else:
                status = "SUCCESS"
            
            cursor.execute("""
                INSERT INTO users (username, email, phone, status, registered_at)
                VALUES (?, ?, ?, ?, datetime('now'))
            """, (username, email, phone, status))
    
    conn.commit()
    conn.close()
    print("✅ นำเข้า CSV → SQLite สำเร็จ")

if __name__ == "__main__":
    excel_to_csv()
    import_to_sqlite()
    
    # ตรวจสอบ
    conn = sqlite3.connect("users.db")
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM users")
    total = cursor.fetchone()[0]
    print(f"📊 รวมทั้งหมด: {total} ราย")
    conn.close()