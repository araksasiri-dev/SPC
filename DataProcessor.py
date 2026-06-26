# DataProcessor.py
import re
from openpyxl import load_workbook

class DataProcessor:
    def __init__(self, excel_path="users_dup.xlsx"):
        self.excel_path = excel_path
        self.raw_data = []
        self.clean_data = []
        self.duplicate_email = []
        self.duplicate_phone = []
    
    def _safe_str(self, value):
        """แปลงค่าเป็น string อย่างปลอดภัย (รองรับ None, int, float)"""
        if value is None:
            return ""
        return str(value).strip()
    
    def read_excel(self):
        """อ่านข้อมูลจาก Excel และแปลงเป็น List of Dictionary"""
        wb = load_workbook(self.excel_path, data_only=True)
        ws = wb.active
        
        # อ่านหัวข้อ (แถวที่ 1)
        headers = [cell.value for cell in ws[1]]
        if not headers or all(h is None for h in headers):
            wb.close()
            return []
        
        # อ่านข้อมูลตั้งแต่แถวที่ 2 เป็นต้นไป
        self.raw_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                break
            user_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    user_dict[header] = row[i]
            self.raw_data.append(user_dict)
        
        wb.close()
        return self.raw_data
    
        
    def process_data(self):
        """ตรวจสอบข้อมูลซ้ำ และแยกข้อมูลเป็นสองกลุ่ม"""
        if not self.raw_data:
            self.read_excel()
        
         # พิมพ์ข้อมูลที่อ่านได้ (debug)
        print("📋 Raw data from Excel:")
        for user in self.raw_data:
            print(f"  {user}")
        
        email_seen = {}
        phone_seen = {}
        
        self.clean_data = []
        self.duplicate_email = []
        self.duplicate_phone = []
        
        for user in self.raw_data:
            email = self._safe_str(user.get("email", ""))
            phone = self._safe_str(user.get("phone", ""))
            
            # ข้ามถ้าอีเมลหรือเบอร์โทรว่าง
            if not email or not phone:
                continue
            
            # ตรวจสอบอีเมลซ้ำ
            if email in email_seen:
                self.duplicate_email.append(user)
                continue
            
            # ตรวจสอบเบอร์โทรซ้ำ
            if phone in phone_seen:
                self.duplicate_phone.append(user)
                continue
            
            # ถ้าผ่านทั้งคู่ ถือว่าข้อมูลดี
            email_seen[email] = True
            phone_seen[phone] = True
            self.clean_data.append(user)
        
        return {
            "clean_data": self.clean_data,
            "duplicate_email": self.duplicate_email,
            "duplicate_phone": self.duplicate_phone
        }
    
    def get_clean_users(self):
        return self.clean_data
    
    def get_duplicate_email_users(self):
        return self.duplicate_email
    
    def get_duplicate_phone_users(self):
        return self.duplicate_phone
    
    def validate_email(self, email):
        """ตรวจสอบรูปแบบอีเมล"""
        email_str = self._safe_str(email)
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email_str) is not None
    
    def write_results_to_excel(self, results, output_path="results_dup.xlsx"):
        wb = load_workbook(self.excel_path)
        ws = wb.active
        
        col_status = ws.max_column + 1
        ws.cell(row=1, column=col_status, value="Status")
        
        # สร้าง mapping username -> status
        status_map = {}
        for item in results:
            username = item.get("username")
            status = item.get("status")
            if username:
                status_map[username] = status
        
        # วนลูปทุกแถวใน Excel
        for row_num in range(2, ws.max_row + 1):
            username_cell = ws.cell(row=row_num, column=1)
            username = username_cell.value
            
            if username in status_map:
                ws.cell(row=row_num, column=col_status, value=status_map[username])
            else:
                ws.cell(row=row_num, column=col_status, value="Skipped (duplicate)")
        
        wb.save(output_path)
        wb.close()