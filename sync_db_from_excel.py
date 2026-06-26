# sync_db_from_excel.py
import sqlite3
from openpyxl import load_workbook

# sync_db_from_excel.py (เวอร์ชันที่มี Duplicate Both)

def sync_database_from_excel(excel_file="users_dup.xlsx"):
    conn = sqlite3.connect("users.db")
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users")
    print("🗑️ ล้าง Database เก่าเรียบร้อย")
    
    wb = load_workbook(excel_file)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"📊 อ่านข้อมูลจาก {excel_file}: {len(rows)} ราย")
    
    success_count = 0
    dup_email_count = 0
    dup_phone_count = 0
    dup_both_count = 0
    email_seen = {}
    phone_seen = {}
    
    for row in rows:
        if not row or not row[0]:
            continue
        
        username = row[0]
        email = row[1] if len(row) > 1 else ""
        phone = row[2] if len(row) > 2 else ""
        
        # ✅ ตรวจสอบทั้ง Email และ Phone
        is_dup_email = email in email_seen
        is_dup_phone = phone in phone_seen
        
        if is_dup_email and is_dup_phone:
            status = "SKIPPED (Duplicate Both)"
            dup_both_count += 1
        elif is_dup_email:
            status = "SKIPPED (Duplicate Email)"
            dup_email_count += 1
        elif is_dup_phone:
            status = "SKIPPED (Duplicate Phone)"
            dup_phone_count += 1
        else:
            status = "SUCCESS"
            success_count += 1
        
        email_seen[email] = True
        phone_seen[phone] = True
        
        cursor.execute("""
            INSERT OR REPLACE INTO users 
            (username, email, phone, status, registered_at)
            VALUES (?, ?, ?, ?, datetime('now'))
        """, (username, email, phone, status))
    
    conn.commit()
    conn.close()
    
    print(f"\n📊 สรุปข้อมูลที่บันทึก:")
    print(f"   ✅ SUCCESS: {success_count} ราย")
    print(f"   📧 SKIPPED (Email): {dup_email_count} ราย")
    print(f"   📱 SKIPPED (Phone): {dup_phone_count} ราย")
    print(f"   ⚠️ SKIPPED (Both): {dup_both_count} ราย")
    print(f"   {'='*40}")
    print(f"   รวมทั้งหมด: {success_count + dup_email_count + dup_phone_count + dup_both_count} ราย")

if __name__ == "__main__":
    sync_database_from_excel("users_dup.xlsx")